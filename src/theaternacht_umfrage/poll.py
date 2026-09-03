"""Build docs/data/poll_results.json from the poll results spreadsheet.

Reads the "Antworten" sheet of ``poll results/poll_results.xlsx`` (written by the
Google Apps Script backend) and joins each vote against ``programm_manual.json``
to attach theatre coordinates, addresses and canonical show times.

Each output entry corresponds to one show that received at least one vote and
carries the list of voters (with their individually chosen time slots). The
route-planning front-end groups entries by ``theaterId`` so that several votes at
the same venue share a single map marker.
"""

from __future__ import annotations

import datetime as _dt
import json
from pathlib import Path

DEFAULT_XLSX = Path("poll results/poll_results.xlsx")
DEFAULT_PROGRAMM = Path("docs/data/programm_manual.json")
DEFAULT_OUT = Path("docs/data/poll_results.json")

SHEET_NAME = "Antworten"


def _norm_time(value: str) -> str:
    """Normalise a single time token to ``HH:MM``."""
    value = value.strip()
    if not value:
        return ""
    # "19:00:00" -> "19:00"
    parts = value.split(":")
    if len(parts) >= 2:
        return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
    return value


def _parse_times(cell: object) -> list[str]:
    """Parse the ``Uhrzeiten`` cell, which may be a comma-joined string or a
    ``datetime.time`` (openpyxl coerces a lone ``HH:MM`` value)."""
    if cell is None:
        return []
    if isinstance(cell, _dt.time):
        return [f"{cell.hour:02d}:{cell.minute:02d}"]
    if isinstance(cell, _dt.datetime):
        return [f"{cell.hour:02d}:{cell.minute:02d}"]
    text = str(cell)
    out: list[str] = []
    for token in text.split(","):
        norm = _norm_time(token)
        if norm:
            out.append(norm)
    return out


def _time_sort_key(t: str) -> tuple[int, int]:
    try:
        h, m = (int(x) for x in t.split(":", 1))
    except ValueError:
        return (99, 99)
    if h < 6:  # after-midnight slots belong at the end of the evening
        h += 24
    return (h, m)


def _load_programm_index(programm_path: Path) -> dict[str, dict]:
    """Map every show id to its theatre + show metadata."""
    data = json.loads(programm_path.read_text(encoding="utf-8"))
    index: dict[str, dict] = {}
    for theater in data.get("theaters", []):
        for show in theater.get("shows", []):
            index[show["id"]] = {
                "theaterId": theater["id"],
                "theaterName": theater["name"],
                "address": theater.get("address", ""),
                "lat": theater.get("lat"),
                "lng": theater.get("lng"),
                "showTitle": show.get("title", ""),
                "venue": show.get("venue", ""),
                "times": list(show.get("times", [])),
            }
    return index


def build_poll(
    out: Path = DEFAULT_OUT,
    xlsx_path: Path = DEFAULT_XLSX,
    programm_path: Path = DEFAULT_PROGRAMM,
) -> dict:
    """Read the spreadsheet, join against the programme and write the JSON."""
    import openpyxl  # local import so the dependency is only needed for --poll

    index = _load_programm_index(programm_path)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet {SHEET_NAME!r} not found in {xlsx_path}")
    ws = wb[SHEET_NAME]

    rows = list(ws.iter_rows(values_only=True))
    # rows[0] is the header: Empfangen, Name, Theater, Programmpunkt, Uhrzeiten,
    # ShowId, Abgesendet
    entries: dict[str, dict] = {}
    voters: set[str] = set()
    unmatched: list[str] = []

    for row in rows[1:]:
        if not row or not row[1]:
            continue
        name = str(row[1]).strip()
        theater_label = str(row[2]).strip() if row[2] else ""
        show_title = str(row[3]).strip() if row[3] else ""
        chosen_times = _parse_times(row[4])
        show_id = str(row[5]).strip() if row[5] else ""
        if not show_id:
            continue
        voters.add(name)

        meta = index.get(show_id)
        if meta is None:
            unmatched.append(show_id)

        entry = entries.get(show_id)
        if entry is None:
            entry = {
                "id": show_id,
                "showId": show_id,
                "theaterId": meta["theaterId"] if meta else "",
                "theaterName": meta["theaterName"] if meta else theater_label,
                "address": meta["address"] if meta else "",
                "lat": meta["lat"] if meta else None,
                "lng": meta["lng"] if meta else None,
                "showTitle": meta["showTitle"] if meta else show_title,
                "venue": meta["venue"] if meta else "",
                "times": meta["times"] if meta else chosen_times,
                "votes": [],
            }
            entries[show_id] = entry

        entry["votes"].append({"voter": name, "times": chosen_times})

    # Deterministic ordering: by theatre name then show title.
    entry_list = sorted(
        entries.values(),
        key=lambda e: (e["theaterName"].lower(), e["showTitle"].lower()),
    )
    for entry in entry_list:
        entry["times"] = sorted(set(entry["times"]), key=_time_sort_key)
        entry["votes"].sort(key=lambda v: v["voter"].lower())

    payload = {
        "generated_at": _dt.datetime.now(_dt.timezone.utc).isoformat(
            timespec="seconds"
        ),
        "source": str(xlsx_path),
        "voters": sorted(voters, key=str.lower),
        "entries": entry_list,
    }

    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )

    if unmatched:
        uniq = sorted(set(unmatched))
        print(
            f"WARNING: {len(uniq)} vote show ids not found in {programm_path}: "
            + ", ".join(uniq)
        )
    return payload
